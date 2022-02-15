import openpyxl
import sqlite3
import numpy as np
from os import listdir

def readTable(filename, file_type):
	#load in excel file and make it active
	wb_obj = openpyxl.load_workbook(filename, data_only=True)
	sheet = wb_obj.active

	#delete rows that are empty at the end of the file
	rowIndex = 0
	for row in sheet.iter_rows():
		if row[0].value == None:
			sheet.delete_rows(rowIndex+2, sheet.max_row-1)
			break
		rowIndex+=1

	if file_type == "pb":
		sheet.delete_cols(sheet.max_column)

	#instantiate variables and create 2D array (a)
	columns = sheet.max_column + 1
	sheet_length = sheet.max_row
	rows = sheet_length-4
	colIndex = 0
	rowIndex = 0
	a = [[0 for x in range(columns)] for y in range(rows)]
	b = [0 for x in range(4)]
	shiftValue = 'A'

	#read through the file and store in table (a)
	for row in sheet.iter_rows(min_row=4, max_row=sheet_length-1):
		for cell in row:
			#adding an extra column that contains shift value (not in spreadsheet)
			if colIndex%columns == 1:
				colIndex+=1
			if cell.value != None:
				if 'Shift' in str(cell.value):
					shiftValue = str(cell.value)[-1]

				# print(rowIndex)
				a[rowIndex][1] = shiftValue

				a[rowIndex][colIndex%columns] = cell.value
			else:
				a[rowIndex][colIndex%columns] = "None"
			colIndex+=1
		rowIndex+=1

	#store all average values into b
	for i in range (3, 7):
		b[i-3] = str(sheet[sheet_length][i].value)

	a = np.array(a)
	b = np.array(b)

	return a, b

def addToDatabase(filepath, a, date, file_type):
	cane_sampling_connection = sqlite3.connect(filepath + 'cane_sampling.sqlite3', check_same_thread=False)
	cane_sampling_cursor = cane_sampling_connection.cursor()

	#basically po# vs no po#
	if file_type == "lab":
		fullCreateString = f'''CREATE TABLE IF NOT EXISTS REPORT{date}(sr INTEGER, shift TEXT, remarks TEXT, token INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE, brix REAL, pol REAL, pty REAL, rec REAL, rec_2 REAL, ded INTEGER);'''
		fullInsertString = f'INSERT OR REPLACE INTO REPORT{date} VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?);'
	else:
		fullCreateString = f'''CREATE TABLE IF NOT EXISTS REPORT{date}(sr INTEGER, shift TEXT, 'pb_no' INTEGER, token INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE, brix REAL, pol REAL, pty REAL, rec REAL, rec_2 REAL, ded INTEGER, remarks TEXT);'''
		fullInsertString = f'INSERT OR REPLACE INTO REPORT{date} VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);'

	#store all values in a into the table
	cane_sampling_cursor.executescript(fullCreateString)
	cane_sampling_cursor.executemany(fullInsertString, a);

	cane_sampling_connection.commit()
	cane_sampling_connection.close()

def averageTable(filepath, b, date):
	cane_sampling_connection = sqlite3.connect(filepath + 'cane_sampling.sqlite3', check_same_thread=False)
	cane_sampling_cursor = cane_sampling_connection.cursor()

	#create average table
	fullCreateString = f'''CREATE TABLE IF NOT EXISTS AVERAGE{date}(brix REAL, pol REAL, pty REAL, rec REAL);'''
	cane_sampling_cursor.executescript(fullCreateString)

	#store values from b in average table
	fullInsertString = f'INSERT OR REPLACE INTO AVERAGE{date} VALUES(?, ?, ?, ?);'
	cane_sampling_cursor.executemany(fullInsertString, [b]);

	cane_sampling_connection.commit()
	cane_sampling_connection.close()


if __name__ == '__main__':
	userDirectory = '/Users/balahaha/Desktop/website-test-copy-master/website/data/databases/'
	directories = ['Excel_data/cane_sampling_lab/', 'Excel_data/cane_sampling_pb/']

	#go through both types of cane sampling files
	for directory in directories:
		#collecting all the cane sampling files
		allFiles = listdir(directory)

		#go through all cane sampling files in this directory and perform functionality
		for file in allFiles:
			if file.endswith(".xlsx") and file.startswith("Cane"):
				#get file type 
				filename = directory + file
				file_type = filename.split("/")[1].split("_")[2]

				#read in table and store in matrices a and b
				[a, b] = readTable(filename, file_type)

				#extract date from excel file name
				date = "_" + filename[-15:-13] + "_" + filename[-12:-10] + "_" + filename[-9:-5]

				#add all info from matrixes a and b into database tables
				addToDatabase(userDirectory, a, date, file_type)
				averageTable(userDirectory, b, date)

